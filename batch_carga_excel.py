#
################################################  START PROCESS  ################################################
#   
import os
import sys
from pathlib import Path
import shutil
import django
from django import setup
from django.db import connection, reset_queries
from openpyxl import load_workbook
from django.db.models import Max

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'inventariohlag.settings')
django.setup()

from tablas.models import *
from inventariohlag.models import *
from inventariohlag.funciones import *

BASE_DIR = Path(__file__).resolve().parent
sys.path.append(str(BASE_DIR))

# === Configurar Django ===
# sys.path.append(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(str(Path(__file__).resolve().parents[1]))


# === Ruta del archivo Excel ===
RUTA_ARCHIVO = os.path.join(".\data", "Gestion_de_Activos_Final.xlsx")  

def elimina_qr():
    qr_dir = Path(BASE_DIR / "media" / "qrcodes")
    if qr_dir.exists() and qr_dir.is_dir():
        print(f"🧹 Eliminando contenido de la carpeta  {qr_dir}...")
        for archivo in qr_dir.iterdir():
            try:
                if archivo.is_file():
                    archivo.unlink()
                elif archivo.is_dir():
                    shutil.rmtree(archivo)
            except Exception as e:
                print(f"⚠️ No se pudo eliminar {archivo}: {e}")
    else:
        print(f"📁 La carpeta {qr_dir} no existe, se creará.")
        qr_dir.mkdir(parents=True, exist_ok=True)
        
        
def siguiente_identificador(tipo):
    maximo = Activos.objects.filter(tipo=tipo).aggregate(Max('newid'))['newid__max']
#    print(connection.queries[-1]['sql'])
    if maximo is None:
        return 1 
    return maximo + 1


def reset_tablas():

    Modelos.objects.all().delete()
    with connection.cursor() as cursor:
        cursor.execute("ALTER TABLE tablas_modelos AUTO_INCREMENT = 1;")
    Fabricantes.objects.all().delete()
    with connection.cursor() as cursor:
        cursor.execute("ALTER TABLE tablas_fabricantes AUTO_INCREMENT = 1;")
    NombresActivos.objects.all().delete()
    with connection.cursor() as cursor:
        cursor.execute("ALTER TABLE tablas_nombresactivos AUTO_INCREMENT = 1;")
    Zonas.objects.all().delete()
    with connection.cursor() as cursor:
        cursor.execute("ALTER TABLE tablas_zonas AUTO_INCREMENT = 1;")
    UsuariosInventario.objects.all().delete()
    Edificios.objects.all().delete()
    with connection.cursor() as cursor:
        cursor.execute("ALTER TABLE tablas_edificios AUTO_INCREMENT = 1;") 
'''    
    with connection.cursor() as cursor:
        cursor.execute("ALTER TABLE tablas_usuariosinventario AUTO_INCREMENT = 1;")
    Paises.objects.all().delete()
    with connection.cursor() as cursor:
        cursor.execute("ALTER TABLE tablas_paises AUTO_INCREMENT = 1;")
    Ciudades.objects.all().delete()
    with connection.cursor() as cursor:
        cursor.execute("ALTER TABLE tablas_ciudades AUTO_INCREMENT = 1;")
'''
    
def importar_excel(hoja):    
    total = 0
    factual = 2
    print(f"Procesando Hoja {hoja.title}.")
    for fila in hoja.iter_rows(min_row=2, max_col=24, values_only=True): 
        atype,aname,modelo,desc1,desc2,desc3,serial,manufacturer,provider,owner,invoice,pdated,pvalue,adated,accounted,uvalue,building,floor,zone,city,country,status,sdated,userinv= fila
        desc1 = str(desc1).strip() if desc1 is not None else ""
        desc2 = str(desc2).strip() if desc2 is not None else ""
        desc3 = str(desc3).strip() if desc3 is not None else ""
        if (is_valid(manufacturer)):
            manufacturer = str(manufacturer).strip()
            try:
                fabricanteid = Fabricantes.objects.get(nombre=manufacturer).id
            except Fabricantes.DoesNotExist:
                fabricanteid = Fabricantes.objects.create(nombre=manufacturer).id
        else:
            fabricanteid = None
        if (is_valid(provider)):
            provider = str(provider).strip()
            try:
                providerid = Proveedores.objects.get(nombre=provider).id
            except Proveedores.DoesNotExist:
                providerid = Proveedores.objects.create(nombre=provider).id
        else:
            providerid = None
        if (is_valid(modelo)):
            modelo = str(modelo).strip()
            try:
                modeloid = Modelos.objects.get(nombre=modelo).id
            except Modelos.DoesNotExist:
                modeloid = Modelos.objects.create(nombre=modelo,fabricante_id=fabricanteid).id
        else:
            modeloid = None
        if (is_valid(aname)):  
            aname = aname.strip()
            try:
                nomactivoid = NombresActivos.objects.get(nombre=aname).id
            except NombresActivos.DoesNotExist:
                nomactivoid = NombresActivos.objects.create(nombre=aname).id
        else:      
            nomactivoid = None
        if (is_valid(zone)):     
            zone = zone.strip()
            try:
                zonaid = Zonas.objects.get(nombre=zone).id
            except Zonas.DoesNotExist:
                zonaid = Zonas.objects.create(nombre=zone).id
        else:     
            zonaid = None
        if (is_valid(userinv)):
            try:
                usuarioid = UsuariosInventario.objects.get(nombre=userinv).id
            except UsuariosInventario.DoesNotExist:
                usuarioid = UsuariosInventario.objects.create(nombre=userinv).id 
        else:
            usuarioid = None
        if (is_valid(country)):     
            country = country.strip()
            try:
                countryid = Paises.objects.get(codigo=country).id
            except Paises.DoesNotExist:
                countryid = Paises.objects.create(nombre=country,codigo=country).id                
        else:     
            countryid = '1'
        if (is_valid(city)):     
            city = city.strip()
            try:
                cityid = Ciudades.objects.get(codigo=city).id
            except Ciudades.DoesNotExist:
                cityid = Ciudades.objects.create(nombre=city,pais_id=countryid,codigo=city).id                  
        else:     
            cityid = '1'
        if (is_valid(building)):     
            building = building.strip()
            try:
                buildingid = Edificios.objects.get(nombre=building).id
            except Edificios.DoesNotExist:
                buildingid = Edificios.objects.create(nombre=building,pais_id=countryid,ciudad_id=cityid).id                   
        else:     
            buildingid = None
        accounted = '1' if is_valid(accounted) else '0'
        if (is_valid(atype)):   
            atype = atype.strip()
            atypeid = next((i for i, (_, nombre) in enumerate(TiposActivos.TIPOS) if nombre == atype), None)
            newid = siguiente_identificador(atypeid)
        else:
            atypeid = 1
        if (is_valid(owner)):  
            owner = owner.strip()
            ownerid = next((i for i, (_, nombre) in enumerate(Owners.OWNERS) if nombre == owner), None)
        else: 
            ownerid = None
        if (is_valid(status)):   
            status = status.strip()
            statusid = next((i for i, (_, nombre) in enumerate(Estados.ESTADOS) if nombre == status), None)
        else:
            statusid = Estados.ACTIV
        adate = fecha_str_to_sql(adated) if adated is not None else fecha_de_hoy_sql()
        pdate = fecha_str_to_sql(pdated) if pdated is not None else fecha_de_hoy_sql()
        sdate = fecha_str_to_sql(sdated) if sdated is not None else fecha_de_hoy_sql()
        try:
            reset_queries()
            Activos.objects.create(id=None,tipo=atypeid,newid=newid,nombre_id=nomactivoid,modelo_id=modeloid,fabricante_id=fabricanteid,desc1=desc1,desc2=desc2,desc3=desc3,
                                    serial=serial,proveedor_id=providerid,owner=ownerid,factura=invoice,fcompra=pdate,vcompra=pvalue,factivacion=adate,
                                    accounted=accounted,vactual=uvalue,building_id=buildingid,floor=floor,zona_id=zonaid,city_id=cityid,country_id=countryid,
                                    estado=statusid,festado=sdate,usuarioinv_id=usuarioid) 
            print(connection.queries[-1]['sql'])
            total += 1        
            print(f"Fila {factual} Ingresada. {fila}")                     
        except Exception as e:
            print(f'Error en Fila {factual} = '+str(e))   
        factual += 1                     
    print(f"En Hoja {hoja.title} Se Ingresaron {total} filas.")


if __name__ == "__main__":
    if not os.path.exists(RUTA_ARCHIVO):
        print(f"Archivo no encontrado: {RUTA_ARCHIVO}")
    else:
#        reset_tablas()
        elimina_qr() 
        Activos.objects.all().delete()
        with connection.cursor() as cursor:
           cursor.execute("ALTER TABLE hlag_activos AUTO_INCREMENT = 1;")
        wb = load_workbook(filename=RUTA_ARCHIVO, data_only=True)
        hoja = wb.worksheets[0] 
        importar_excel(hoja)
        hoja = wb.worksheets[1] 
        importar_excel(hoja)

