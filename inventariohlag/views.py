from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import get_object_or_404, render,redirect
from django.http import HttpResponse
from django.core import exceptions
from django.contrib.auth import logout
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from inventariohlag.funciones import *
from .models import *
from tablas.models import *
from django.db.models import Q


def login(request):
    next_url = request.GET.get('next') or request.POST.get('next')
    if (next_url is None):
        next_url = ""
    return render(request, 'login.html',{'next_url': next_url})

@csrf_exempt
def salir_plataforma(request):
    '''
    Realiza logout y envia home 
    '''
    logout(request)
    return redirect('login')

def home_admin(request):
    if not request.user.is_superuser:
        return render(request, '404.html')
    return render(request, 'home_admin.html')

def home(request):
    return render(request,'home.html')

def about(request):
    return render(request,'about.html')

def ayuda(request):
    return render(request,'ayuda.html')

def test(request):
    return render(request, 'test.html')

def en_desarrollo(request):
    return render(request, '204.html')


@login_required(login_url='/login')
def activos_filtrar(request,info):
     
    if (request.method == 'GET'):
        persona = request.user.persona
        if (info == 0):
            titulo = "Assets Report"
            
            accion = 'activos_listar'
        else:
            titulo = "QR Codes Report"
            accion = 'qr_listar'
        areas = NombresAreas.objects.all().order_by('nombre')
        paises = Paises.objects.filter(
            id__in=Areas.objects.filter(areaname_id=persona.area_id).values_list('pais_id', flat=True)
            ).order_by('nombre')
        ciudades =  Ciudades.objects.filter(pais_id=persona.pais_id).order_by('nombre')
        if (request.user.is_superuser):
            if (persona.tipoactivo == TiposActivos.TODOS):
                tiposactivos = TiposActivos.get_choices(exclude=TiposActivos.TODOS)
            else:
                tiposactivos = TiposActivos.get_choices(include=persona.tipoactivo)
        else:
            tiposactivos = TiposActivos.get_choices(include=persona.tipoactivo)
#        print(str(activos.query))
        return render(request,'activos_filtrar.html', {'persona': persona, 'tiposactivos': tiposactivos, 'areas': areas, 'paises': paises, 'ciudades': ciudades, 
                                                       'accion': accion, 'titulo': titulo, 'info': info})
    else:
        return render(request, '404.html')    


@login_required(login_url='/login')
def activos_listar(request):
#    if not request.user.is_superuser:
#        return render(request, '404.html')
    if (request.method == 'POST'):
        persona = request.user.persona
        owners = Owners.OWNERS
        contabilizados = Accounted.ACCOUNTED
        estados = Estados.ESTADOS
        edificios =  Edificios.objects.all().order_by('nombre')
        ciudades = Ciudades.objects.all().order_by('nombre')
        paises = Paises.objects.all().order_by('nombre')
        proveedores = Proveedores.objects.all().order_by('nombre')
        modelos = Modelos.objects.all().order_by('nombre')
        fabricantes = Fabricantes.objects.all().order_by('nombre')
        nombresactivos = NombresActivos.objects.all().order_by('nombre')
        zonas = Zonas.objects.all().order_by('nombre')
        usuariosinv = UsuariosInventario.objects.all().order_by('nombre')
        tipo = request.POST.get('tipo')
        ciudad = request.POST.get('ciudad')
        filtro = Q(tipo=tipo) & Q(city_id=ciudad)
        activos = Activos.objects.filter(filtro).order_by('tipo','newid')
        tiposactivos = TiposActivos.get_choices(include=tipo)
#       print(str(activos.query))
        return render(request,'activos_listar.html', {'activos': activos,'owners': owners,'edificios': edificios,'ciudades': ciudades,'paises': paises,
                                                    'tiposactivos': tiposactivos,'estados': estados,'modelos': modelos,
                                                    'fabricantes': fabricantes,'nombresactivos': nombresactivos,'proveedores': proveedores,'zonas': zonas,
                                                    'usuariosinv': usuariosinv,'contabilizados': contabilizados})
    else:
        return render(request, '404.html')    


@login_required(login_url='/login')
def checklastid(request):
    context = {}
    context['status'] = 500
    context['newid'] = 0
    
    if (request.method == 'POST'):
        tipo = request.POST.get('tipo')
        context['status'] = 200
        context['newid'] = siguiente_identificador(tipo)
    return JsonResponse({
        'context': context,
        })


def siguiente_identificador(tipo):
    from django.db.models import Max
    maximo = Activos.objects.filter(tipo=tipo).aggregate(Max('newid'))['newid__max']
    if maximo is None:
        return 1   # si no hay registros, empezar en 1
    return maximo + 1


@login_required(login_url='/login')
def leeractivo(request):
    context = {}
    context['status'] = 200
    context['message'] = ""
    registro = {}
    if (request.method == 'POST'):
        id = request.POST.get('id')
        try:
            activo = Activos.objects.get(id=id)
        except Activos.DoesNotExist:
            context['status'] = 404  
            context['message'] = f'Error: Asset Not Available'
        except Exception as e: 
            context['status'] = 404 
            context['message'] = f'Error: {e}'
        else:
            registro['id'] = activo.id
            registro['tipo'] = activo.tipo
            registro['newid'] = activo.newid            
            registro['nombre'] = activo.nombre_id
            registro['modelo'] = activo.modelo_id
            registro['fabricante'] = activo.fabricante_id
            registro['sku'] = activo.sku
            registro['desc1'] = activo.desc1
            registro['desc2'] = activo.desc2
            registro['desc3'] = activo.desc3
            registro['serial'] = activo.serial
            registro['proveedor'] = activo.proveedor_id
            registro['owner'] = activo.owner
            registro['factura'] = activo.factura
            registro['fcompra'] = fecha_sql_to_str(activo.fcompra)            
            registro['vcompra'] = activo.vcompra
            registro['factivacion'] = fecha_sql_to_str(activo.factivacion)
            registro['accounted'] = activo.accounted
            registro['vactual'] = activo.vactual
            registro['building'] = activo.building_id
            registro['floor'] = activo.floor
            registro['zona'] = activo.zona_id
            registro['city'] = activo.city_id
            registro['country'] = activo.country_id
            registro['estado'] = activo.estado
            registro['festado'] = fecha_sql_to_str(activo.festado)
            registro['usuarioinv'] = activo.usuarioinv_id
            registro['fingreso'] = activo.fecha_hora_fingreso
            registro['fmodifica'] = activo.fecha_hora_fmodifica                  
    else:
        context['status'] = 404 
        context['message'] = f'Access Error'
#    print(registro,context) 
    return JsonResponse({
        'context': context,
        'registro': registro,
        })


@login_required(login_url='/login')
def modificaractivo(request):
    if not request.user.is_superuser:
        return render(request, '404.html')
    if (request.method == 'POST'):
        accion = request.POST.get('accion')
        id = request.POST.get('id')
        tipo = request.POST.get('tipo')        
        nombre = request.POST.get('nombre')
        modelo = request.POST.get('modelo')
        fabricante = request.POST.get('fabricante')
        sku = request.POST.get('sku')
        desc1 = request.POST.get('desc1')
        desc2 = request.POST.get('desc2')
        desc3 = request.POST.get('desc3')
        serial = request.POST.get('serial')
        proveedor = request.POST.get('proveedor')
        owner = request.POST.get('owner')
        factura = request.POST.get('factura')
        fcompra = fecha_str_to_sql(request.POST.get('fcompra'))           
        vcompra = request.POST.get('vcompra')
        factivacion = fecha_str_to_sql(request.POST.get('factivacion'))
        accounted = request.POST.get('accounted')
        vactual = request.POST.get('vactual')
        building = request.POST.get('building')
        floor = request.POST.get('floor')
        zona = request.POST.get('zona')
        city = request.POST.get('city')
        country = request.POST.get('country')
        estado = request.POST.get('estado')
        festado = fecha_str_to_sql(request.POST.get('festado'))
        usuarioinv = request.POST.get('usuarioinv')
        if (accion == AccionesCrud.ELIMINAR):
            try:
                Activos.objects.filter(id=id).delete()
            except Exception as e:
                print('Accion= ' + accion + ' Error= '+str(e))
        elif (accion == AccionesCrud.CREAR):
            try:
                Activos.objects.create(tipo=tipo,nombre_id=nombre,modelo_id=modelo,fabricante_id=fabricante,
                                       desc1=desc1,desc2=desc2,desc3=desc3,serial=serial,proveedor_id=proveedor,owner=owner,factura=factura,fcompra=fcompra,
                                       vcompra=vcompra,factivacion=factivacion,accounted=accounted,vactual=vactual,
                                       building_id=building,floor=floor,zona_id=zona,city_id=city,country_id=country,sku=sku,
                                       estado=estado,festado=festado,usuarioinv_id=usuarioinv)
            except Exception as e:
                print('Accion= ' + accion + ' Error= '+str(e))
        else:
            try:
                Activos.objects.filter(id=id).update(tipo=tipo,nombre_id=nombre,modelo_id=modelo,fabricante_id=fabricante,
                                       desc1=desc1,desc2=desc2,desc3=desc3,serial=serial,proveedor_id=proveedor,owner=owner,factura=factura,fcompra=fcompra,
                                       vcompra=vcompra,factivacion=factivacion,accounted=accounted,vactual=vactual,
                                       building_id=building,floor=floor,zona_id=zona,city_id=city,country_id=country,sku=sku,
                                       estado=estado,festado=festado,usuarioinv_id=usuarioinv)
            except Exception as e:
                print('Accion= ' + accion + ' Error= '+str(e))                
        persona = request.user.persona
        if (request.user.is_superuser):
            if (persona.tipoactivo == TiposActivos.TODOS):
                tipos_activos = [t[0] for t in TiposActivos.get_choices(exclude=TiposActivos.TODOS)]
            else:
                tipos_activos = [t[0] for t in TiposActivos.get_choices(include=persona.tipoactivo)]
            filtro = Q(tipo__in=tipos_activos)
            activos = Activos.objects.filter(filtro).order_by('tipo','newid')
        else:
            paises_areas = Paises.objects.filter(areas__id=persona.area_id)
            filtro = Q(country__in=paises_areas)
            activos = Activos.objects.filter(filtro,tipo=persona.tipoactivo).order_by('tipo','newid')           
        return render(request, 'activos_ajax_listar.html', {'activos': activos})
    else:
        return render(request, '404.html')    


@login_required(login_url='/login')
def qr_detalle(request, token):
    from urllib.parse import unquote
    from django.core import signing
    from django.core.signing import SignatureExpired, BadSignature

    decoded = unquote(token).rstrip('/')
    try:
        # Decodifica el token (no expira)
        data = signing.loads(decoded)
        id = data['id']
        activo = get_object_or_404(Activos, id=id)
    except SignatureExpired:
        return HttpResponse("QR Code Expired. Please try scanning a new code.", status=403)
    except BadSignature:
        return HttpResponse("Invalid QR Code.", status=403)
    owners = Owners.OWNERS
    contabilizados = Accounted.ACCOUNTED
    estados = Estados.ESTADOS
    tiposactivos = TiposActivos.get_choices(exclude=TiposActivos.TODOS)
    return render(request, 'qr_detalle_activo.html', {'activo': activo,'owners': owners,'tiposactivos': tiposactivos,'contabilizados': contabilizados,'estados': estados})


def qr_detalle_test(request):

    id = request.GET.get('id')
    activo = get_object_or_404(Activos, id=id)
    owner = Owners.OWNERS[int(activo.owner)][1]
    contabilizados = Accounted.ACCOUNTED
    estados = Estados.ESTADOS
    tiposactivos = TiposActivos.get_choices(exclude=TiposActivos.TODOS)
    return render(request, 'qr_detalle_activo.html', {'activo': activo,'owner': owner,'tiposactivos': tiposactivos,'contabilizados': contabilizados,'estados': estados})


@login_required(login_url='/login')
def ver_qr(request):
    context = {}
    context['status'] = 500
    context['qr_url'] = ""
    context['message'] = ""
    
    if (request.method == 'POST'):
        id = request.POST.get('id')
        try:
            activo = Activos.objects.get(id=id)
        except Exception as e:
            context['message'] = f"{e}"
        else:
#            tipo = TiposActivos.TIPOS[int(activo.tipo)][1]
            context['status'] = 200
            context['qr_url'] = f"{settings.SITE_URL}{settings.MEDIA_URL}{activo.qr}"
#        print(context)
    return JsonResponse({
        'context': context,
        })


@login_required(login_url='/login')
def qr_listar(request):
     
    if (request.method == 'POST'):
        tipo = request.POST.get('tipo')
        pais = request.POST.get('pais')
        ciudad = request.POST.get('ciudad')
        frow = request.POST.get('frow')
        fcol = request.POST.get('fcol')
        try: 
            frow = int(frow)
        except: 
            frow = 1
        if not (1 <= frow <= 8):
            frow = 1
        try: 
            fcol = int(fcol)
        except:
            fcol = 1
        if not (1 <= fcol <= 3):
            fcol = 1
        fcell = 3 * (frow -  1 ) + fcol    
        tipo = request.POST.get('tipo')
        ciudad = request.POST.get('ciudad')
        filtro = Q(tipo=tipo) & Q(city_id=ciudad)
        activos = Activos.objects.filter(filtro).order_by('tipo','newid')
#        print(str(activos.query))
        return render(request,'qr_listar.html', {'activos': activos, 'fcell': fcell})
    else:
        return render(request, '404.html')    


@login_required(login_url='/login')

def areas_listar(request):
    if (request.method == 'GET'):
        areasnombres = NombresAreas.objects.all().order_by('nombre')
        areas = Areas.objects.all().order_by('areaname_id')
        paises = Paises.objects.all().order_by('nombre')
        return render(request, 'areas_listar.html', {'areas': areas, 'paises': paises, 'areasnombres': areasnombres})
    else: 
        accion = request.POST.get('accion')
        id = request.POST.get('id')
        pais = request.POST.get('pais')
        areaname = request.POST.get('areaname')        
        if (accion == AccionesCrud.ELIMINAR):
            try:
                Areas.objects.filter(id=id).delete()
            except Exception as e:
                print('Accion= ' + accion + ' Error= '+str(e))
        elif (accion == AccionesCrud.CREAR):
            try:
                Areas.objects.create(pais_id=pais, areaname_id=areaname)
            except Exception as e:
                print('Accion= ' + accion + ' Error= '+str(e))
        else:
            try:
                Areas.objects.filter(id=id).update(pais_id=pais, areaname_id=areaname)
            except Exception as e:
                print('Accion= ' + accion + ' Error= '+str(e))
        areas = Areas.objects.all().order_by('areaname_id')
        return render(request, 'areas_ajax_listar.html', {'areas': areas})
    
def excel_master(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Protection,  Alignment, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Formulario"

    # Fila de titulos
    headers = ["Active Type", "Active Name", "Model", "Descriptive 1","Descriptive 2","Descriptive 3","Serial Number","Manufacturer","Provider","Owner","Invoice Number",
               "Purchase Date","Purchase Value","Activation Date","Accounted","Update Value","Building","Floor","Zone","Location","Country","Status","Status Date","User"]
    for col_num, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_num, value=header)
#    for col in range(1, len(headers) + 1):
#        cell = ws.cell(row=1, column=col)
#        cell.font = Font(bold=True)
    fill_gris = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    for cell in ws[1]:
        cell.fill = fill_gris
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    # 1. Obtener opciones de Active Name    
    opciones  = list(NombresActivos.objects.values_list('nombre', flat=True))
    # 2. Crear hoja auxiliar
    ws_list = wb.create_sheet(title="nactivos")
    for i, val in enumerate(opciones , start=1):
        ws_list[f"A{i}"] = val
    # 3. Crear validación de datos (lista)
    rango = f"nactivos!$A$1:$A${len(opciones )}"
    dv = DataValidation(type="list", formula1=rango, allow_blank=False)
    ws.add_data_validation(dv)
    # 4. Aplicar a un rango
    dv.add("B2:B100")
    # 5. Ocultar hoja de Active Name 
    ws_list.sheet_state = "hidden"

    # 1. Obtener opciones de Active Model    
    opciones  = list(Modelos.objects.values_list('nombre', flat=True))
    # 2. Crear hoja auxiliar
    ws_list = wb.create_sheet(title="modelac")
    for i, val in enumerate(opciones , start=1):
        ws_list[f"A{i}"] = val
    # 3. Crear validación de datos (lista)
    rango = f"modelac!$A$1:$A${len(opciones )}"
    dv = DataValidation(type="list", formula1=rango, allow_blank=False)
    ws.add_data_validation(dv)
    # 4. Aplicar a un rango
    dv.add("C2:C100")
    # 5. Ocultar hoja de Active Model 
    ws_list.sheet_state = "hidden"
            
    # 1. Obtener opciones de Fabricantes
    opciones  = list(Fabricantes.objects.values_list('nombre', flat=True))
    # 2. Crear hoja auxiliar
    ws_list = wb.create_sheet(title="fabricantes")
    for i, val in enumerate(opciones , start=1):
        ws_list[f"A{i}"] = val
    # 3. Crear validación de datos (lista)
    rango = f"fabricantes!$A$1:$A${len(opciones )}"
    dv = DataValidation(type="list", formula1=rango, allow_blank=False)
    ws.add_data_validation(dv)
    # 4. Aplicar a un rango (ej: columna B filas 2 a 100)
    dv.add("H2:H100")
    # 5. Ocultar hoja de fabricantes
    ws_list.sheet_state = "hidden"

    # 1. Obtener opciones de Proveedores
    opciones  = list(Proveedores.objects.values_list('nombre', flat=True))
    # 2. Crear hoja auxiliar
    ws_list = wb.create_sheet(title="proveedores")
    for i, val in enumerate(opciones , start=1):
        ws_list[f"A{i}"] = val
    # 3. Crear validación de datos (lista)
    rango = f"proveedores!$A$1:$A${len(opciones )}"
    dv = DataValidation(type="list", formula1=rango, allow_blank=False)
    ws.add_data_validation(dv)
    # 4. Aplicar a un rango (ej: columna B filas 2 a 100)
    dv.add("I2:I100")
    # 5. Ocultar hoja de proveedores
    ws_list.sheet_state = "hidden"
    
    # 1. Obtener opciones de Edificios
    opciones  = list(Edificios.objects.values_list('nombre', flat=True))
    # 2. Crear hoja auxiliar
    ws_list = wb.create_sheet(title="edificios")
    for i, val in enumerate(opciones , start=1):
        ws_list[f"A{i}"] = val
    # 3. Crear validación de datos (lista)
    rango = f"edificios!$A$1:$A${len(opciones )}"
    dv = DataValidation(type="list", formula1=rango, allow_blank=False)
    ws.add_data_validation(dv)
    # 4. Aplicar a un rango (ej: columna B filas 2 a 100)
    dv.add("Q2:Q100")
    # 5. Ocultar hoja de edificios
    ws_list.sheet_state = "hidden"
    
    # 1. Obtener opciones de Ciudades
    opciones  = list(Ciudades.objects.values_list('nombre', flat=True))
    # 2. Crear hoja auxiliar
    ws_list = wb.create_sheet(title="ciudades")
    for i, val in enumerate(opciones , start=1):
        ws_list[f"A{i}"] = val
    # 3. Crear validación de datos (lista)
    rango = f"ciudades!$A$1:$A${len(opciones )}"
    dv = DataValidation(type="list", formula1=rango, allow_blank=False)
    ws.add_data_validation(dv)
    # 4. Aplicar a un rango (ej: columna B filas 2 a 100)
    dv.add("T2:T100")
    # 5. Ocultar hoja de ciudades
    ws_list.sheet_state = "hidden"
    
    # 1. Obtener opciones de Paises
    opciones  = list(Paises.objects.values_list('nombre', flat=True))
    # 2. Crear hoja auxiliar
    ws_list = wb.create_sheet(title="paises")
    for i, val in enumerate(opciones , start=1):
        ws_list[f"A{i}"] = val
    # 3. Crear validación de datos (lista)
    rango = f"paises!$A$1:$A${len(opciones )}"
    dv = DataValidation(type="list", formula1=rango, allow_blank=False)
    ws.add_data_validation(dv)
    # 4. Aplicar a un rango (ej: columna B filas 2 a 100)
    dv.add("U2:U100")
    # 5. Ocultar hoja de paises
    ws_list.sheet_state = "hidden"
    
    # 1. Obtener opciones de Usuarios Inventario
    opciones  = list(UsuariosInventario.objects.values_list('nombre', flat=True))
    # 2. Crear hoja auxiliar
    ws_list = wb.create_sheet(title="user_assets")
    for i, val in enumerate(opciones , start=1):
        ws_list[f"A{i}"] = val
    # 3. Crear validación de datos (lista)
    rango = f"user_assets!$A$1:$A${len(opciones )}"
    dv = DataValidation(type="list", formula1=rango, allow_blank=False)
    ws.add_data_validation(dv)
    # 4. Aplicar a un rango (ej: columna B filas 2 a 100)
    dv.add("X2:X100")
    # 5. Ocultar hoja de user_assets
    ws_list.sheet_state = "hidden"  
    
        
    MIN_WIDTH = 20
    MAX_WIDTH = 60
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max(min(max_length + 2, MAX_WIDTH), MIN_WIDTH)
        ws.column_dimensions[col_letter].width = adjusted_width

    ws.freeze_panes = "A1"
    for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=24):
        for cell in row:
            cell.protection = Protection(locked=True)
    for row in ws.iter_rows(min_row=2, max_row=100, min_col=1, max_col=24):
        for cell in row:
            cell.protection = Protection(locked=False)   
            
    # 6. (Opcional) proteger hoja
    ws.protection.sheet = True

    # 7. Respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    response['Content-Disposition'] = 'attachment; filename="formulario.xlsx"'

    wb.save(response)
    return response


def excel_upload(request):
    return render(request, '204.html')
